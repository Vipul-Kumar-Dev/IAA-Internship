from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.shortcuts import render
from django.contrib import messages
import requests
from .models import Faculty, Infrastructure, Course, Catering
from django.contrib.auth import authenticate, login, logout, get_user_model

User = get_user_model()

from openpyxl import Workbook
from django.http import HttpResponse

def welcome_page(request):
    return render(request, 'welcome.html')

def download_faculty_feedback(request):
    faculty_feedbacks = Faculty.objects.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=faculty_feedbacks.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Feedbacks"
    
    # Create header row
    headers = ['Faculty Name', 'Satisfaction', 'Behavior', 'Knowledge', 'Interaction', 'Clarity', 'Response', 'Examples', 'Motivation', 'Comments']
    ws.append(headers)
    
    # Add feedback data
    for feedback in faculty_feedbacks:
        ws.append([feedback.faculty_name, feedback.satisfaction, feedback.behavior, feedback.knowledge, feedback.interaction, feedback.clarity, feedback.response, feedback.examples, feedback.motivation, feedback.description])
    
    wb.save(response)
    return response

def download_infrastructure_feedback(request):
    infrastructure_feedbacks = Infrastructure.objects.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=infrastructure_feedbacks.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Infrastructure Feedbacks"
    
    # Create header row
    headers = ['Infrastructure Name', 'Satisfaction', 'Quality', 'Resources', 'Maintenance', 'Safety', 'Comments']
    ws.append(headers)
    
    # Add feedback data
    for feedback in infrastructure_feedbacks:
        ws.append([feedback.infrastructure_name, feedback.satisfaction, feedback.quality, feedback.resources, feedback.maintenance, feedback.safety, feedback.description])
    
    wb.save(response)
    return response

def download_course_feedback(request):
    course_feedbacks = Course.objects.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=course_feedbacks.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Feedbacks"
    
    # Create header row
    headers = ['Course Name', 'Satisfaction', 'Content', 'Instructor', 'Materials', 'Comments']
    ws.append(headers)
    
    # Add feedback data
    for feedback in course_feedbacks:
        ws.append([feedback.course_name, feedback.satisfaction, feedback.content, feedback.instructor, feedback.materials, feedback.description])
    
    wb.save(response)
    return response

def download_catering_feedback(request):
    catering_feedbacks = Catering.objects.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=catering_feedbacks.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Catering Feedbacks"
    
    # Create header row
    headers = ['Catering Name', 'Overall Satisfaction', 'Food Quality', 'Service Quality', 'Cleanliness', 'Affordable', 'Comments']
    ws.append(headers)
    
    # Add feedback data
    for feedback in catering_feedbacks:
        ws.append([feedback.catering_name, feedback.overall_satisfaction, feedback.food_quality, feedback.service_quality, feedback.Cleanliness, feedback.Affordable, feedback.description])
    
    wb.save(response)
    return response

    return render(request, "welcome.html")

def login_page(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                request.session['user_firstname'] = user.first_name
                request.session['user_id'] = user.id

                if user.is_superuser:
                    return redirect('adminhome')
                else:
                    return redirect('home')

            else:
                messages.error(request, "Your account is inactive.")
                return redirect('login')
        else:
            messages.error(request, "Invalid username or password.")
            return redirect('login')

    return render(request, "login.html")

EMAIL_VERIFICATION_API_KEY = "dc1b4e7595fd4049a2f39e315775177c"
EMAIL_VERIFICATION_URL = "https://emailvalidation.abstractapi.com/v1/?api_key=dc1b4e7595fd4049a2f39e315775177c&email=vipulsam1234@gmail.com/"

def verify_email(email):
    response = requests.get(f"{EMAIL_VERIFICATION_URL}?api_key={EMAIL_VERIFICATION_API_KEY}&email={email}")
    if response.status_code == 200:
        data = response.json()
        return data.get("deliverability") == "DELIVERABLE"  # Check if the email exists
    return False

def signup(request):
    if request.method == "POST":
        firstname = request.POST.get('firstname')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if not all([firstname, email, username, password, confirm_password]):
            messages.error(request, "All fields are required.")
            return redirect('signup')

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('signup')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already registered.")
            return redirect('signup')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username is already taken.")
            return redirect('signup')

        if not verify_email(email):
            messages.error(request, "Please enter a valid email address.")
            return redirect('signup')

        user = User.objects.create_user(
            first_name=firstname,
            email=email,
            username=username,
            password=password
        )
        user.is_active = True
        user.save()
        messages.success(request, "Signup successful! You can now log in.")
        return redirect('login')

    return render(request, "login.html")

@login_required
def home_page(request):
    user_id = request.session.get('user_id')
    if not user_id:
        messages.error(request, "You must log in to access the home page.")
        return redirect('login')

    firstname = request.session.get('user_firstname', '')
    return render(request, "home.html", {"firstname": firstname, "messages": messages.get_messages(request)})


def logout_view(request):
    if not request.user.is_staff:
        logout(request)
        request.session.flush()
        messages.success(request, "You have successfully logged out.")
    else:
        messages.success(request, "You have successfully logged out.")
    
    return redirect('login')

@login_required
def facultyfeed(request):
    if request.method == "POST":
        faculty_name = request.POST.get("faculty")
        behavior = request.POST.get("behavior")
        knowledge = request.POST.get("knowledge")
        interaction = request.POST.get("interaction")
        clarity = request.POST.get("clarity")
        response = request.POST.get("response")
        examples = request.POST.get("examples")
        motivation = request.POST.get("motivation")
        satisfaction = request.POST.get("satisfaction")
        description = request.POST.get("comments", "")

        # Validate numeric fields
        try:
            behavior = int(behavior)
            knowledge = int(knowledge)
            interaction = int(interaction)
            clarity = int(clarity)
            response = int(response)
            examples = int(examples)
            motivation = int(motivation)
            satisfaction = int(satisfaction)

            # Ensure values are within range
            for rating in [behavior, knowledge, interaction, clarity, response, examples, motivation, satisfaction]:
                if rating < 1 or rating > 5:
                    messages.error(request, "Each rating must be between 1 and 5.")
                    return redirect('facultyfeed')

        except ValueError:
            messages.error(request, "Invalid rating value. Please enter a number between 1 and 5.")
            return redirect('facultyfeed')

        # Fetch the user instance
        try:
            user_instance = User.objects.get(id=request.session.get("user_id"))
        except User.DoesNotExist:
            messages.error(request, "User not found. Please log in again.")
            return redirect('login')

        # Save the feedback
        Faculty.objects.create(
            trainee=user_instance,
            faculty_name=faculty_name,
            behavior=behavior,
            knowledge=knowledge,
            interaction=interaction,
            clarity=clarity,
            response=response,
            examples=examples,
            motivation=motivation,
            satisfaction=satisfaction,
            description=description
        )

        messages.success(request, "Feedback submitted successfully!")
        return redirect('thankyou')

    return render(request, "facultyfeed.html")

@login_required
def infrafeed(request):
    if request.method == "POST":
        infrastructure_name = request.POST.get("infrastructure")
        quality = request.POST.get("quality")
        resources = request.POST.get("resources")
        maintenance = request.POST.get("maintenance")
        safety = request.POST.get("safety")
        satisfaction = request.POST.get("satisfaction")
        description = request.POST.get("comments", "")

        try:
            quality = int(quality)
            resources = int(resources)
            maintenance = int(maintenance)
            safety = int(safety)
            satisfaction = int(satisfaction)

            for rating in [quality, resources, maintenance, safety, satisfaction]:
                if rating < 1 or rating > 5:
                    messages.error(request, "Each rating must be between 1 and 5.")
                    return redirect('infrafeed')

        except ValueError:
            messages.error(request, "Invalid rating value. Please enter a number between 1 and 5.")
            return redirect('infrafeed')

        try:
            user_instance = User.objects.get(id=request.session.get("user_id"))
        except User.DoesNotExist:
            messages.error(request, "User not found. Please log in again.")
            return redirect('login')

        # Save the feedback
        Infrastructure.objects.create(
            trainee=user_instance,
            infrastructure_name=infrastructure_name,
            quality=quality,
            resources=resources,
            maintenance=maintenance,
            safety=safety,
            satisfaction=satisfaction,
            description=description
        )

        messages.success(request, "Feedback submitted successfully!")
        return redirect('thankyou')

    return render(request, "infrafeed.html")

@login_required
def coursefeed(request):
    if request.method == "POST":
        course_name = request.POST.get("course_name")
        content = request.POST.get("content")
        instructor = request.POST.get("instructor")
        materials = request.POST.get("materials")
        satisfaction = request.POST.get("satisfaction")
        description = request.POST.get("comments", "")

        try:
            content = int(content)
            instructor = int(instructor)
            materials = int(materials)
            satisfaction = int(satisfaction)

            for rating in [content, instructor, materials, satisfaction]:
                if rating < 1 or rating > 5:
                    messages.error(request, "Each rating must be between 1 and 5.")
                    return redirect('infrafeed')

        except ValueError:
            messages.error(request, "Invalid rating value. Please enter a number between 1 and 5.")
            return redirect('infrafeed')

        try:
            user_instance = User.objects.get(id=request.session.get("user_id"))
        except User.DoesNotExist:
            messages.error(request, "User not found. Please log in again.")
            return redirect('login')

        # Save the feedback
        Course.objects.create(
            trainee=user_instance,
            course_name=course_name,
            content=content,
            instructor=instructor,
            materials=materials,
            satisfaction=satisfaction,
            description=description
        )

        messages.success(request, "Feedback submitted successfully!")
        return redirect('thankyou')

    return render(request, "coursefeed.html")

@login_required
def cateringfeed(request):
    if request.method == "POST":
        catering_name = request.POST.get("catering_name")
        food_quality = request.POST.get("food_quality")
        service_quality = request.POST.get("service_quality")
        Cleanliness = request.POST.get("Cleanliness")
        Affordable = request.POST.get("Affordable")
        overall_satisfaction = request.POST.get("overall_satisfaction")
        description = request.POST.get("comments", "")

        try:
            food_quality = int(food_quality)
            service_quality = int(service_quality)
            Cleanliness = int(Cleanliness)
            Affordable = int(Affordable)
            overall_satisfaction = int(overall_satisfaction)

            for rating in [food_quality, service_quality, Cleanliness, Affordable, overall_satisfaction]:
                if rating < 1 or rating > 5:
                    messages.error(request, "Each rating must be between 1 and 5.")
                    return redirect('infrafeed')

        except ValueError:
            messages.error(request, "Invalid rating value. Please enter a number between 1 and 5.")
            return redirect('infrafeed')

        try:
            user_instance = User.objects.get(id=request.session.get("user_id"))
        except User.DoesNotExist:
            messages.error(request, "User not found. Please log in again.")
            return redirect('login')

        # Save the feedback
        Catering.objects.create(
            trainee=user_instance,
            catering_name=catering_name,
            food_quality=food_quality,
            service_quality=service_quality,
            Cleanliness=Cleanliness,
            Affordable=Affordable,
            overall_satisfaction=overall_satisfaction,
            description=description
        )

        messages.success(request, "Feedback submitted successfully!")
        return redirect('thankyou')

    return render(request, "cateringfeed.html")

def contact(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        user_email = request.POST.get("email")
        message = request.POST.get("message")

        if full_name and user_email and message:
            subject = f"New Contact Form Submission from {full_name}"
            message_body = f"Name: {full_name}\nEmail: {user_email}\nMessage: {message}"
            from_email = "vipulsam1234@gmail.com"
            recipient_email = ["anas.shaikh7827@gmail.com"]

            try:
                email_message = EmailMessage(
                    subject=subject,
                    body=message_body,
                    from_email=from_email,
                    to=recipient_email,
                    reply_to=[user_email],
                )
                email_message.send()
                messages.success(request, "Your message has been sent successfully!")
            except Exception as e:
                messages.error(request, "Error sending message. Please try again later.")

        return redirect("contact")

    return render(request, "contact.html")

@login_required
def admin_home(request):
    faculty_feedbacks = Faculty.objects.all()
    infrastructure_feedbacks = Infrastructure.objects.all()
    course_feedbacks = Course.objects.all()
    catering_feedbacks = Catering.objects.all()
    
    context = {
        'faculty_feedbacks': faculty_feedbacks,
        'infrastructure_feedbacks': infrastructure_feedbacks,
        'course_feedbacks': course_feedbacks,
        'catering_feedbacks': catering_feedbacks,
    }
    return render(request, 'adminhome.html', context)


def thankyou(request):
    return render(request, "thankyou.html")

def serve_meta(request):
    return JsonResponse({
        "status": "ok",
        "message": "This is a placeholder meta.json file"
    })
